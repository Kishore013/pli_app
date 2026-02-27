"""
PLI / RPLI vs APT Comparison — Flask Backend
=============================================
• Receives parsed row data (JSON) from the browser (XLSX.js does the Excel parsing)
• Stores everything in SQLite  (pli_comparison.db)
• Primary keys:
    pli / rpli  →  receipt_number
    apt         →  (posting_date, office_id, gl_code, total_amount)
• Comparison is done in SQL for speed and accuracy
"""

from flask import Flask, request, jsonify, send_file
from contextlib import contextmanager
import sqlite3, json, os, io, csv, datetime, re

app = Flask(__name__)

# Allow requests from any origin (needed if HTML is ever opened as a local file)
@app.after_request
def add_cors(response):
    response.headers["Access-Control-Allow-Origin"]  = "*"
    response.headers["Access-Control-Allow-Headers"] = "Content-Type"
    response.headers["Access-Control-Allow-Methods"] = "GET, POST, DELETE, OPTIONS"
    return response

@app.route("/api/<path:p>", methods=["OPTIONS"])
def options_handler(p):
    return "", 204

DB_PATH = os.path.join(os.path.dirname(__file__), "pli_comparison.db")

# ─────────────────────────────────────────────
#  DB CONNECTION — always closes after use
# ─────────────────────────────────────────────
@contextmanager
def get_db():
    """
    Open a connection, yield it, commit on success, rollback on error, ALWAYS close.
    timeout=30 means SQLite waits up to 30 s for a write lock before raising.
    check_same_thread=False is safe here because we close immediately after each use.
    """
    conn = sqlite3.connect(DB_PATH, timeout=30, check_same_thread=False)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA foreign_keys = ON")
    try:
        yield conn
        conn.commit()
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()   # ← critical: release the file lock immediately

# ─────────────────────────────────────────────
#  DB INIT  (WAL set once here, not per-request)
# ─────────────────────────────────────────────
def init_db():
    # Set WAL mode first in its own connection so it persists to the DB file
    _conn = sqlite3.connect(DB_PATH, timeout=30)
    _conn.execute("PRAGMA journal_mode = WAL")
    _conn.execute("PRAGMA synchronous  = NORMAL")   # safe with WAL, faster than FULL
    _conn.commit()
    _conn.close()

    with get_db() as conn:
        conn.executescript("""
            CREATE TABLE IF NOT EXISTS master (
                mcc_name     TEXT NOT NULL,
                apt_name     TEXT NOT NULL,
                apt_code     TEXT NOT NULL,
                office_type  TEXT,
                PRIMARY KEY (mcc_name, apt_code)
            );

            CREATE TABLE IF NOT EXISTS acc_classification (
                trans_id      TEXT,
                category      TEXT NOT NULL,
                sub_category  TEXT,
                account_head  TEXT,
                account_code  TEXT NOT NULL,
                PRIMARY KEY (account_code)
            );

            CREATE TABLE IF NOT EXISTS pli (
                receipt_number  TEXT PRIMARY KEY,
                office_name     TEXT,
                apt_code        TEXT,
                effective_date  TEXT,
                gross_amount    REAL,
                is_bodi         INTEGER DEFAULT 0
            );

            CREATE TABLE IF NOT EXISTS rpli (
                receipt_number  TEXT PRIMARY KEY,
                office_name     TEXT,
                apt_code        TEXT,
                effective_date  TEXT,
                gross_amount    REAL
            );

            CREATE TABLE IF NOT EXISTS apt (
                posting_date   TEXT NOT NULL,
                office_id      TEXT NOT NULL,
                gl_code        TEXT NOT NULL,
                total_amount   REAL NOT NULL,
                credit_debit   TEXT,
                signed_amount  REAL,
                PRIMARY KEY (posting_date, office_id, gl_code, total_amount)
            );

            CREATE TABLE IF NOT EXISTS upload_log (
                id          INTEGER PRIMARY KEY AUTOINCREMENT,
                table_name  TEXT NOT NULL,
                rows_loaded INTEGER,
                uploaded_at TEXT DEFAULT (datetime('now','localtime'))
            );

            CREATE TABLE IF NOT EXISTS settlements (
                report_type  TEXT NOT NULL,
                apt_code     TEXT NOT NULL,
                eff_date     TEXT NOT NULL,
                mcc_amount   REAL,
                apt_amount   REAL,
                difference   REAL,
                remark       TEXT,
                settled_at   TEXT DEFAULT (datetime('now','localtime')),
                PRIMARY KEY (report_type, apt_code, eff_date)
            );
        """)
    print(f"[DB] Initialised at {DB_PATH}")

init_db()

# ─────────────────────────────────────────────
#  HELPERS
# ─────────────────────────────────────────────
def normalize_receipt(val):
    if val is None: return ""
    s = str(val).strip().replace(" ", "")
    s = re.sub(r'\.0+$', '', s)
    s = s.lstrip('0')
    return s

def safe_float(val):
    if val is None or val == "": return 0.0
    try:
        return float(str(val).replace(",", ""))
    except:
        return 0.0

def _log(conn, table_name, rows):
    """Write upload log row using the SAME open connection — no second lock."""
    conn.execute(
        "INSERT INTO upload_log (table_name, rows_loaded) VALUES (?,?)",
        (table_name, rows)
    )

# ─────────────────────────────────────────────
#  ROUTES — PAGES
# ─────────────────────────────────────────────
@app.route("/")
def index():
    html_path = os.path.join(os.path.dirname(__file__), "templates", "index.html")
    if not os.path.exists(html_path):
        # Also check same directory as app.py (in case templates/ folder missing)
        html_path = os.path.join(os.path.dirname(__file__), "index.html")
    with open(html_path, "r", encoding="utf-8") as f:
        return f.read(), 200, {"Content-Type": "text/html; charset=utf-8"}

# ─────────────────────────────────────────────
#  ROUTE — SAVE MASTER
# ─────────────────────────────────────────────
@app.route("/api/save/master", methods=["POST"])
def save_master():
    rows = request.json.get("rows", [])
    if not rows:
        return jsonify({"error": "No rows received"}), 400

    saved = 0
    with get_db() as conn:
        conn.execute("DELETE FROM master")
        for r in rows:
            mcc  = str(r.get("mcc_name","")).strip().upper()
            name = str(r.get("apt_name","")).strip()
            code = str(r.get("apt_code","")).strip()
            otype= str(r.get("office_type","")).strip().upper()
            if not mcc or not name or not code: continue
            conn.execute(
                "INSERT OR REPLACE INTO master VALUES (?,?,?,?)",
                (mcc, name, code, otype or None)
            )
            saved += 1
        _log(conn, "master", saved)
    return jsonify({"saved": saved})

# ─────────────────────────────────────────────
#  ROUTE — SAVE ACC CLASSIFICATION
# ─────────────────────────────────────────────
@app.route("/api/save/acc_classification", methods=["POST"])
def save_acc_classification():
    rows = request.json.get("rows", [])
    if not rows:
        return jsonify({"error": "No rows received"}), 400

    saved = 0
    with get_db() as conn:
        conn.execute("DELETE FROM acc_classification")
        for r in rows:
            trans    = str(r.get("trans_id","")).strip()
            cat      = str(r.get("category","")).strip().upper()
            subcat   = str(r.get("sub_category","")).strip()
            head     = str(r.get("account_head","")).strip()
            code     = str(r.get("account_code","")).strip()
            if not cat or not code: continue
            conn.execute(
                "INSERT OR REPLACE INTO acc_classification VALUES (?,?,?,?,?)",
                (trans or None, cat, subcat or None, head or None, code)
            )
            saved += 1
        _log(conn, "acc_classification", saved)
    return jsonify({"saved": saved})

# ─────────────────────────────────────────────
#  ROUTE — SAVE APT
# ─────────────────────────────────────────────
@app.route("/api/save/apt", methods=["POST"])
def save_apt():
    rows = request.json.get("rows", [])
    if not rows:
        return jsonify({"error": "No rows received"}), 400

    saved = 0
    with get_db() as conn:
        conn.execute("DELETE FROM apt")
        for r in rows:
            pdate  = str(r.get("posting_date","")).strip()
            oid    = str(r.get("office_id","")).strip()
            gl     = str(r.get("gl_code","")).strip()
            amt    = safe_float(r.get("total_amount", 0))
            crdr   = str(r.get("credit_debit","")).strip().upper()
            signed = -amt if crdr == "P" else amt
            if not pdate or not oid or not gl: continue
            conn.execute(
                """INSERT OR REPLACE INTO apt
                   (posting_date, office_id, gl_code, total_amount, credit_debit, signed_amount)
                   VALUES (?,?,?,?,?,?)""",
                (pdate, oid, gl, amt, crdr or None, signed)
            )
            saved += 1
        _log(conn, "apt", saved)
    return jsonify({"saved": saved})

# ─────────────────────────────────────────────
#  ROUTE — SAVE PLI
# ─────────────────────────────────────────────
@app.route("/api/save/pli", methods=["POST"])
def save_pli():
    rows       = request.json.get("rows", [])
    bodi_codes = set(request.json.get("bodi_receipts", []))
    if not rows:
        return jsonify({"error": "No rows received"}), 400

    added = 0
    skipped = 0
    with get_db() as conn:
        # Pull master map from DB
        master_rows = conn.execute("SELECT mcc_name, apt_code FROM master").fetchall()
        master_map  = {}
        for mr in master_rows:
            mcc = mr["mcc_name"]
            if mcc not in master_map:
                master_map[mcc] = []
            master_map[mcc].append(mr["apt_code"])

        for r in rows:
            office  = str(r.get("office_name","")).strip().upper()
            receipt = normalize_receipt(r.get("receipt_number",""))
            date    = str(r.get("effective_date","")).strip()
            amount  = safe_float(r.get("gross_amount", 0))
            if not office or not receipt or not date: continue

            codes = master_map.get(office)
            if not codes: continue

            # BODI split
            if office == "BODINAICKENPATTI" and len(codes) >= 2:
                apt_code = "29103265" if receipt in bodi_codes else "29103266"
            else:
                apt_code = codes[0]

            is_bodi = 1 if office == "BODINAICKENPATTI" else 0

            cur = conn.execute(
                """INSERT OR IGNORE INTO pli
                   (receipt_number, office_name, apt_code, effective_date, gross_amount, is_bodi)
                   VALUES (?,?,?,?,?,?)""",
                (receipt, office, apt_code, date, amount, is_bodi)
            )
            if cur.rowcount:
                added += 1
            else:
                skipped += 1

        _log(conn, "pli", added)
    return jsonify({"saved": added, "added": added, "skipped": skipped})

# ─────────────────────────────────────────────
#  ROUTE — SAVE RPLI
# ─────────────────────────────────────────────
@app.route("/api/save/rpli", methods=["POST"])
def save_rpli():
    rows       = request.json.get("rows", [])
    bodi_codes = set(request.json.get("bodi_receipts", []))
    if not rows:
        return jsonify({"error": "No rows received"}), 400

    added = 0
    skipped = 0
    with get_db() as conn:
        master_rows = conn.execute("SELECT mcc_name, apt_code FROM master").fetchall()
        master_map  = {}
        for mr in master_rows:
            mcc = mr["mcc_name"]
            if mcc not in master_map:
                master_map[mcc] = []
            master_map[mcc].append(mr["apt_code"])

        for r in rows:
            office  = str(r.get("office_name","")).strip().upper()
            receipt = normalize_receipt(r.get("receipt_number",""))
            date    = str(r.get("effective_date","")).strip()
            amount  = safe_float(r.get("gross_amount", 0))
            if not office or not receipt or not date: continue

            codes = master_map.get(office)
            if not codes: continue

            if office == "BODINAICKENPATTI" and len(codes) >= 2:
                apt_code = "29103265" if receipt in bodi_codes else "29103266"
            else:
                apt_code = codes[0]

            cur = conn.execute(
                """INSERT OR IGNORE INTO rpli
                   (receipt_number, office_name, apt_code, effective_date, gross_amount)
                   VALUES (?,?,?,?,?)""",
                (receipt, office, apt_code, date, amount)
            )
            if cur.rowcount:
                added += 1
            else:
                skipped += 1

        _log(conn, "rpli", added)
    return jsonify({"saved": added, "added": added, "skipped": skipped})

# ─────────────────────────────────────────────
#  ROUTE — COMPARE
# ─────────────────────────────────────────────
@app.route("/api/compare", methods=["POST"])
def compare():
    body        = request.json or {}
    report_type = body.get("report_type", "PLI").upper()   # PLI or RPLI
    date_filter = body.get("date_filter", "")              # YYYY-MM-DD
    month_filter= body.get("month_filter", "")             # YYYY-MM
    office_type = body.get("office_type", "ALL").upper()
    diff_only   = body.get("diff_only", False)

    src_table = "pli" if report_type == "PLI" else "rpli"

    with get_db() as conn:
        # Aggregate McCamish amounts: apt_code + date → sum(gross_amount)
        mcc_sql = f"""
            SELECT apt_code, effective_date, SUM(gross_amount) AS mcc_total
            FROM {src_table}
            GROUP BY apt_code, effective_date
        """
        mcc_rows = {
            (r["apt_code"], r["effective_date"]): r["mcc_total"]
            for r in conn.execute(mcc_sql).fetchall()
        }

        # Aggregate APT amounts filtered by GL Code classification
        apt_sql = """
            SELECT a.office_id, a.posting_date, SUM(a.signed_amount) AS apt_total
            FROM apt a
            JOIN acc_classification ac ON ac.account_code = a.gl_code
            WHERE ac.category = ?
            GROUP BY a.office_id, a.posting_date
        """
        apt_rows = {
            (r["office_id"], r["posting_date"]): r["apt_total"]
            for r in conn.execute(apt_sql, (report_type,)).fetchall()
        }

        # Union of all keys
        all_keys = set(mcc_rows.keys()) | set(apt_rows.keys())

        # Load settled keys to exclude from comparison
        settled_keys = set(
            (r["apt_code"], r["eff_date"])
            for r in conn.execute(
                "SELECT apt_code, eff_date FROM settlements WHERE report_type = ?",
                (report_type,)
            ).fetchall()
        )

        # Pull office metadata
        meta = {
            r["apt_code"]: {"name": r["apt_name"], "type": r["office_type"] or ""}
            for r in conn.execute("SELECT apt_code, apt_name, office_type FROM master").fetchall()
        }

        results = []
        for (apt_code, date) in sorted(all_keys):
            # Skip settled entries
            if (apt_code, date) in settled_keys:
                continue

            # Date filter
            if month_filter and not date.startswith(month_filter):
                continue
            if date_filter and date != date_filter:
                continue

            # Office type filter
            if office_type != "ALL":
                if meta.get(apt_code, {}).get("type","").upper() != office_type:
                    continue

            mcc_amt = mcc_rows.get((apt_code, date), 0.0)
            apt_amt = apt_rows.get((apt_code, date), 0.0)
            diff    = round(mcc_amt - apt_amt, 2)

            if diff_only and diff == 0:
                continue

            results.append({
                "apt_code":   apt_code,
                "apt_name":   meta.get(apt_code, {}).get("name", apt_code),
                "date":       date,
                "mcc_amount": round(mcc_amt, 2),
                "apt_amount": round(apt_amt, 2),
                "difference": diff
            })

        # Totals
        total_mcc  = round(sum(r["mcc_amount"] for r in results), 2)
        total_apt  = round(sum(r["apt_amount"]  for r in results), 2)
        total_diff = round(sum(r["difference"]  for r in results), 2)
        offices    = len(set(r["apt_code"] for r in results))
        mismatches = sum(1 for r in results if r["difference"] != 0)

    return jsonify({
        "report_type": report_type,
        "rows":        results,
        "totals": {
            "mcc_total":  total_mcc,
            "apt_total":  total_apt,
            "difference": total_diff,
            "offices":    offices,
            "mismatches": mismatches
        }
    })

# ─────────────────────────────────────────────
#  ROUTE — DB STATUS (table row counts + last upload)
# ─────────────────────────────────────────────
@app.route("/api/db/status")
def db_status():
    with get_db() as conn:
        tables = ["master", "acc_classification", "pli", "rpli", "apt"]
        status = {}
        for t in tables:
            count = conn.execute(f"SELECT COUNT(*) as c FROM {t}").fetchone()["c"]
            last  = conn.execute(
                "SELECT uploaded_at, rows_loaded FROM upload_log WHERE table_name=? ORDER BY id DESC LIMIT 1", (t,)
            ).fetchone()
            status[t] = {
                "rows":        count,
                "last_upload": last["uploaded_at"] if last else None,
                "last_rows":   last["rows_loaded"] if last else None
            }
    return jsonify(status)

# ─────────────────────────────────────────────
#  ROUTE — VIEW TABLE DATA (paginated)
# ─────────────────────────────────────────────
@app.route("/api/db/table/<table_name>")
def view_table(table_name):
    allowed = ["master", "acc_classification", "pli", "rpli", "apt", "upload_log"]
    if table_name not in allowed:
        return jsonify({"error": "Table not allowed"}), 403

    page     = int(request.args.get("page", 1))
    per_page = int(request.args.get("per_page", 100))
    offset   = (page - 1) * per_page

    with get_db() as conn:
        total = conn.execute(f"SELECT COUNT(*) as c FROM {table_name}").fetchone()["c"]
        rows  = conn.execute(
            f"SELECT * FROM {table_name} LIMIT ? OFFSET ?", (per_page, offset)
        ).fetchall()
        cols  = [d[0] for d in conn.execute(f"SELECT * FROM {table_name} LIMIT 0").description or []]

    return jsonify({
        "table":    table_name,
        "total":    total,
        "page":     page,
        "per_page": per_page,
        "columns":  cols,
        "rows":     [dict(r) for r in rows]
    })

# ─────────────────────────────────────────────
#  ROUTE — DELETE TABLE
# ─────────────────────────────────────────────
@app.route("/api/db/delete/<table_name>", methods=["DELETE"])
def delete_table(table_name):
    allowed = ["master", "acc_classification", "pli", "rpli", "apt"]
    if table_name not in allowed:
        return jsonify({"error": "Table not allowed"}), 403

    with get_db() as conn:
        conn.execute(f"DELETE FROM {table_name}")
        conn.execute(
            "INSERT INTO upload_log (table_name, rows_loaded) VALUES (?, 0)",
            (f"{table_name}_cleared",)
        )
    return jsonify({"cleared": table_name})

# ─────────────────────────────────────────────
#  ROUTE — EXPORT DB FILE
# ─────────────────────────────────────────────
@app.route("/api/db/export")
def export_db():
    if not os.path.exists(DB_PATH):
        return jsonify({"error": "DB not found"}), 404
    return send_file(
        DB_PATH,
        as_attachment=True,
        download_name=f"pli_comparison_{datetime.date.today().isoformat()}.db",
        mimetype="application/octet-stream"
    )

# ─────────────────────────────────────────────
#  ROUTE — EXPORT COMPARISON AS CSV
# ─────────────────────────────────────────────
@app.route("/api/export/csv", methods=["POST"])
def export_csv():
    body = request.json or {}
    report_type = body.get("report_type", "BOTH").upper()

    output = io.StringIO()
    writer = csv.writer(output)

    types = ["PLI", "RPLI"] if report_type == "BOTH" else [report_type]

    with get_db() as conn:
        meta = {
            r["apt_code"]: {"name": r["apt_name"], "type": r["office_type"] or ""}
            for r in conn.execute("SELECT apt_code, apt_name, office_type FROM master").fetchall()
        }

        # Build settled key sets for each report type
        settled_keys_by_type = {}
        for rtype in types:
            settled_keys_by_type[rtype] = set(
                (r["apt_code"], r["eff_date"])
                for r in conn.execute(
                    "SELECT apt_code, eff_date FROM settlements WHERE report_type=?", (rtype,)
                ).fetchall()
            )

        for rtype in types:
            writer.writerow([f"=== {rtype} DIFFERENCES ==="])
            writer.writerow(["S.No", "Office Code", "Office Name", "Date",
                             f"McCamish ({rtype})", "APT Amount", "Difference"])

            src = "pli" if rtype == "PLI" else "rpli"
            mcc_rows = {
                (r["apt_code"], r["effective_date"]): r["mcc_total"]
                for r in conn.execute(
                    f"SELECT apt_code, effective_date, SUM(gross_amount) AS mcc_total FROM {src} GROUP BY apt_code, effective_date"
                ).fetchall()
            }
            apt_rows = {
                (r["office_id"], r["posting_date"]): r["apt_total"]
                for r in conn.execute(
                    """SELECT a.office_id, a.posting_date, SUM(a.signed_amount) AS apt_total
                       FROM apt a JOIN acc_classification ac ON ac.account_code = a.gl_code
                       WHERE ac.category = ? GROUP BY a.office_id, a.posting_date""",
                    (rtype,)
                ).fetchall()
            }
            all_keys = set(mcc_rows.keys()) | set(apt_rows.keys())

            sno = 1
            for (code, date) in sorted(all_keys):
                # Skip settled entries in CSV export too
                if (code, date) in settled_keys_by_type.get(rtype, set()):
                    continue
                mcc_amt = mcc_rows.get((code, date), 0.0)
                apt_amt = apt_rows.get((code, date), 0.0)
                diff    = round(mcc_amt - apt_amt, 2)
                if diff == 0: continue
                writer.writerow([
                    sno, code,
                    meta.get(code, {}).get("name", code),
                    date,
                    round(mcc_amt, 2), round(apt_amt, 2), diff
                ])
                sno += 1
            writer.writerow([])  # blank line between sections

    output.seek(0)
    return send_file(
        io.BytesIO(output.getvalue().encode("utf-8-sig")),
        as_attachment=True,
        download_name=f"PLI_RPLI_Differences_{datetime.date.today().isoformat()}.csv",
        mimetype="text/csv"
    )


# ─────────────────────────────────────────────
#  ROUTE — SAVE SETTLEMENT (upsert)
# ─────────────────────────────────────────────
@app.route("/api/settlements/save", methods=["POST"])
def save_settlement():
    d = request.json or {}
    report_type = d.get("report_type","").upper()
    apt_code    = d.get("apt_code","").strip()
    eff_date    = d.get("eff_date","").strip()
    remark      = d.get("remark","").strip()
    mcc_amount  = float(d.get("mcc_amount", 0))
    apt_amount  = float(d.get("apt_amount", 0))
    difference  = float(d.get("difference", 0))

    if not report_type or not apt_code or not eff_date:
        return jsonify({"error": "Missing required fields"}), 400

    with get_db() as conn:
        conn.execute("""
            INSERT INTO settlements
                (report_type, apt_code, eff_date, mcc_amount, apt_amount, difference, remark, settled_at)
            VALUES (?,?,?,?,?,?,?, datetime('now','localtime'))
            ON CONFLICT(report_type, apt_code, eff_date)
            DO UPDATE SET remark=excluded.remark,
                          mcc_amount=excluded.mcc_amount,
                          apt_amount=excluded.apt_amount,
                          difference=excluded.difference,
                          settled_at=datetime('now','localtime')
        """, (report_type, apt_code, eff_date, mcc_amount, apt_amount, difference, remark))

    return jsonify({"ok": True})

# ─────────────────────────────────────────────
#  ROUTE — REMOVE SETTLEMENT (unsettle)
# ─────────────────────────────────────────────
@app.route("/api/settlements/remove", methods=["POST"])
def remove_settlement():
    d = request.json or {}
    report_type = d.get("report_type","").upper()
    apt_code    = d.get("apt_code","").strip()
    eff_date    = d.get("eff_date","").strip()
    if not report_type or not apt_code or not eff_date:
        return jsonify({"error": "Missing required fields"}), 400
    with get_db() as conn:
        conn.execute(
            "DELETE FROM settlements WHERE report_type=? AND apt_code=? AND eff_date=?",
            (report_type, apt_code, eff_date)
        )
    return jsonify({"ok": True})

# ─────────────────────────────────────────────
#  ROUTE — LIST SETTLEMENTS
# ─────────────────────────────────────────────
@app.route("/api/settlements/list", methods=["POST"])
def list_settlements():
    d = request.json or {}
    report_type = d.get("report_type","BOTH").upper()
    with get_db() as conn:
        meta = {
            r["apt_code"]: r["apt_name"]
            for r in conn.execute("SELECT apt_code, apt_name FROM master").fetchall()
        }
        if report_type == "BOTH":
            rows = conn.execute("SELECT * FROM settlements ORDER BY report_type, eff_date").fetchall()
        else:
            rows = conn.execute(
                "SELECT * FROM settlements WHERE report_type=? ORDER BY eff_date", (report_type,)
            ).fetchall()
    result = [{
        "report_type": r["report_type"],
        "apt_code":    r["apt_code"],
        "apt_name":    meta.get(r["apt_code"], r["apt_code"]),
        "eff_date":    r["eff_date"],
        "mcc_amount":  r["mcc_amount"],
        "apt_amount":  r["apt_amount"],
        "difference":  r["difference"],
        "remark":      r["remark"] or "",
        "settled_at":  r["settled_at"]
    } for r in rows]
    return jsonify({"rows": result, "total": len(result)})

# ─────────────────────────────────────────────
#  ROUTE — EXPORT SETTLED ENTRIES AS XLSX
# ─────────────────────────────────────────────
@app.route("/api/settlements/export", methods=["POST"])
def export_settlements():
    d = request.json or {}
    report_type = d.get("report_type","BOTH").upper()
    with get_db() as conn:
        meta = {
            r["apt_code"]: r["apt_name"]
            for r in conn.execute("SELECT apt_code, apt_name FROM master").fetchall()
        }
        if report_type == "BOTH":
            rows = conn.execute("SELECT * FROM settlements ORDER BY report_type, eff_date").fetchall()
        else:
            rows = conn.execute(
                "SELECT * FROM settlements WHERE report_type=? ORDER BY eff_date", (report_type,)
            ).fetchall()
    if not rows:
        return jsonify({"error": "No settled entries found"}), 404

    # Build XLSX in-memory using only stdlib csv as fallback if openpyxl unavailable
    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment
        from openpyxl.utils import get_column_letter
        wb = openpyxl.Workbook()
        types_present = list(dict.fromkeys(r["report_type"] for r in rows))
        for idx, rtype in enumerate(types_present):
            ws = wb.active if idx == 0 else wb.create_sheet()
            ws.title = f"{rtype} Settled"
            headers = ["S.No","Report Type","Office Code","Office Name","Date",
                       f"McCamish ({rtype})","APT Amount","Difference","Remark","Settled At"]
            for col, h in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=h)
                cell.font      = Font(bold=True, color="0A0C11")
                cell.fill      = PatternFill("solid", fgColor="00E5A0")
                cell.alignment = Alignment(horizontal="center")
            type_rows = [r for r in rows if r["report_type"] == rtype]
            for sno, r in enumerate(type_rows, 1):
                row_data = [sno, r["report_type"], r["apt_code"],
                            meta.get(r["apt_code"], r["apt_code"]),
                            r["eff_date"], r["mcc_amount"], r["apt_amount"],
                            r["difference"], r["remark"] or "", r["settled_at"]]
                for col, val in enumerate(row_data, 1):
                    cell = ws.cell(row=sno+1, column=col, value=val)
                    if r["difference"] != 0:
                        cell.fill = PatternFill("solid", fgColor="2D1520")
            for col, w in enumerate([6,12,14,30,12,18,18,18,40,20], 1):
                ws.column_dimensions[get_column_letter(col)].width = w
            ws.freeze_panes = "A2"
        buf = io.BytesIO()
        wb.save(buf); buf.seek(0)
        return send_file(buf, as_attachment=True,
            download_name=f"Settled_Entries_{report_type}_{datetime.date.today().isoformat()}.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    except ImportError:
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(["S.No","Report Type","Office Code","Office Name","Date",
                         "McCamish Amount","APT Amount","Difference","Remark","Settled At"])
        for sno, r in enumerate(rows, 1):
            writer.writerow([sno, r["report_type"], r["apt_code"],
                meta.get(r["apt_code"], r["apt_code"]),
                r["eff_date"], r["mcc_amount"], r["apt_amount"],
                r["difference"], r["remark"] or "", r["settled_at"]])
        output.seek(0)
        return send_file(io.BytesIO(output.getvalue().encode("utf-8-sig")), as_attachment=True,
            download_name=f"Settled_Entries_{report_type}_{datetime.date.today().isoformat()}.csv",
            mimetype="text/csv")

if __name__ == "__main__":
    print("=" * 55)
    print("  PLI/RPLI Comparison Server — starting…")
    print()
    print("  ✅ Open your browser and go to:")
    print("     http://localhost:5000")
    print()
    print("  ⚠️  Do NOT open index.html directly as a file.")
    print("     It must be served through this server.")
    print("=" * 55)
    #app.run(
    #    debug=False,          # disables Werkzeug debugger + PIN lock entirely
    #    port=5000,
    #    threaded=True,
    #    use_reloader=False
    #)
    
    app.run(host='0.0.0.0', port=5000)
